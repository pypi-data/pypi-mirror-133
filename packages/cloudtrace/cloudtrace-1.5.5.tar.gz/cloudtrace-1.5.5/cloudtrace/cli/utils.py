import os
import subprocess
from abc import ABC, abstractmethod
from multiprocessing.pool import ThreadPool
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook

def write(rows, outfile, sheet_name=None, **kwargs):
    df = pd.DataFrame(rows)
    for k, v in kwargs.items():
        df[k] = v
    df.sort_values(by='name', inplace=True)
    exists = os.path.exists(outfile)
    if exists:
        try:
            book = load_workbook(outfile)
            try:
                del book[sheet_name]
            except KeyError:
                pass
        except BadZipFile:
            exists = False
    writer = pd.ExcelWriter(outfile, engine='openpyxl')
    if exists:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

def get_hostname(args):
    addr, user = args
    cmd = 'ssh -o StrictHostKeyChecking=no {}@{} hostname'.format(user, addr)
    p = subprocess.Popen(cmd, shell=True, universal_newlines=True, stdout=subprocess.PIPE)
    outs, _ = p.communicate()
    outs = outs.strip()
    return addr, outs

class AbstractCLI(ABC):
    def __init__(self, country=None):
        self.country = country

    def get_hostnames(self, addrs, user):
        addrs = [(addr, user) for addr in addrs]
        hostnames = {}
        with ThreadPool(25) as pool:
            for addr, hostname in pool.imap_unordered(get_hostname, addrs):
                hostnames[addr] = hostname
        return hostnames

    @abstractmethod
    def public_ip_addresses(self, info=None):
        raise NotImplementedError()

    @abstractmethod
    def get_region_names(self):
        raise NotImplementedError()

    def run_all(self, func):
        regnames = self.get_region_names()
        with ThreadPool(min(50, len(regnames))) as pool:
            for _ in pool.imap_unordered(func, regnames):
                pass

    @abstractmethod
    def start_region(self, args):
        raise NotImplementedError()

    @abstractmethod
    def state_region(self, args):
        raise NotImplementedError()

    @abstractmethod
    def stop_region(self, args):
        raise NotImplementedError()

    def write(self, outfile, sheet_name, user, **kwargs):
        rows = []
        self.public_ip_addresses(info=rows)
        df = pd.DataFrame(rows)
        df['user'] = user
        for k, v in kwargs.items():
            df[k] = v
        hostnames = self.get_hostnames(df.addr, user)
        df['hostname'] = df.addr.map(hostnames)
        df.sort_values(by='name', inplace=True)
        exists = os.path.exists(outfile)
        if exists:
            try:
                book = load_workbook(outfile)
                try:
                    del book[sheet_name]
                except KeyError:
                    pass
            except BadZipFile:
                exists = False
        writer = pd.ExcelWriter(outfile, engine='openpyxl')
        if exists:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()