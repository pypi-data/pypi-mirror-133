'''
This class gathers data from excel in a 2d array
Author: Robyn Proffer July 2018
'''


'''TODO

'''

#used libraries and modules
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter



class DataBender:

    def Retrieve(self, file, mini = 0, maxi = None, colbreak = True):
        """
        Retrieve data from within a workbook.
        File is the name of the workbook to pull from, mini is the minimum row which defaults to the topmost row, maxi
        is the final row to process and defaults to the final row, colbreak is a boolean which determines
        whether or not the processing of a row ends on a blank column, which defaults to yes.
        """
        sheet = []
        inBook = openpyxl.load_workbook(filename = file)
        dataSheet = inBook.active #could do getsheet here too
        count = 0
        if maxi == None:
            maxi = dataSheet.max_row
        header = list(dataSheet.rows)[0]
        for row in dataSheet.iter_rows(min_row = mini, max_row = maxi):
            newrow = []
            for col in row:
                if col.value == None:
                    if colbreak == True:
                        break
                newrow.append(str(col.value))
            sheet.append(newrow)
        inBook.close
        return sheet
    
    def putOut(self, data, title, heading, Freeze = False, color = None): #title need be a filename prepped for saving. EG: terminal.xlsx. data must be a two dimensional array
        outBook = openpyxl.Workbook()
        outSheet = outBook.active      
        outSheet.append(heading) #heading needs to be a 1d array
        bfont = Font(bold=True)
        for cell in outSheet["1:1"]:
            cell.font = bfont                 
        for a in data: #figure out the structure underlying data
            outSheet.append(a)
        dims = {}
        for row in outSheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            outSheet.column_dimensions[col].width = value + 5       
        if color != None:
            self.colorformat(outSheet, color, "yellow")
        if Freeze != False:
            outSheet.freeze_panes = "B2"
        outBook.save(title)
        outBook.close
        
    def as_text(value):
        if value is None:
            return ""
        return str(value)    
        
    def putOutMultiBook(self, data, title, heading, Freeze = False, color = None): #title need be a filename prepped for saving. EG: terminal.xlsx. data must be a dictionary of sheetnames to 2D arrays
        outBook = openpyxl.Workbook()
        for sheet in data:
            if color is not None:
                cb = color[sheet]
            outSheet = outBook.create_sheet(sheet)
            if isinstance(heading, dict):
                outSheet.append(heading[sheet]) #heading needs to be a dictionary of page titles to headings
            if isinstance(heading, list):
                outSheet.append(heading)
            bfont = Font(bold=True)
            for cell in outSheet["1:1"]:
                cell.font = bfont     
            for a in data[sheet]: #figure out the structure underlying data
                outSheet.append(a)
            dims = {}
            for row in outSheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                outSheet.column_dimensions[col].width = value + 5        
            if color != None:
                self.colorformat(outSheet, cb, "yellow")
            if Freeze != False:
                outSheet.freeze_panes = "B2"
        try:
            outBook.remove(outBook["Sheet"])
        except:
            print("named sheet not present")
        finally:    
            outBook.save(title)
            outBook.close 
        
    '''
    this is almost certainly wonky
    '''
    
    def colorformat(sheet, cords, color): ##cords is a dictionary of row to list of columns to highlight. color is the color.
        color = openpyxl.styles.colors.Color(rgb='00FF0000')     
        #fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color)      
        for r in cords:
            colorrow = cords[r]
            for i in colorrow:
                #print(r)
                #print(i)
                c = sheet.cell(row = r, column = i)
                c.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
    
    def staticformat():
        pass
