import openpyxl
from xlsxhelper import get_workbook
from fastutils import strutils

from django.utils.translation import ugettext_lazy as _

DEFAULT_FILENAME_TEMPLATE = "{model_name}-{year:04d}{month:02d}{day:02d}-{hour:02d}{minute:02d}{second:02d}.xlsx"
XLSX_CELL_SIZE_MAX = 32*1024


class Aggregate(object):
    pass

class Sum(Aggregate):
    def __init__(self):
        self.total = 0
    
    def push(self, value):
        self.total += value
    
    def final(self):
        return self.total

class Average(Aggregate):
    def __init__(self, empty="-"):
        self.empty = empty
        self.total = 0
        self.count = 0

    def push(self, value):
        self.total += value
        self.count += 1
    
    def final(self):
        if not self.count:
            return self.empty
        else:
            return self.total / self.count

class Count(Aggregate):
    def __init__(self, value=0):
        self.value = value
        self.count = 0
    
    def push(self, value):
        if self.value == value:
            self.count += 1
    
    def final(self):
        return self.count


class ForceStringRender(object):
    def __call__(self, value):
        return str(value)

class DateRender(object):
    def __init__(self, format="%Y/%m/%d", empty_value="-"):
        self.format = format
        self.empty_value = empty_value

    def __call__(self, value):
        if not value:
            return self.empty_value
        else:
            return value.strftime(self.format)

class BooleanRender(object):
    def __init__(self, true_display="TRUE", false_display="FALSE"):
        self.true_display = true_display
        self.false_display = false_display
    
    def __call__(self, value):
        if value:
            return self.true_display
        else:
            return self.false_display

class NullBooleanRender(object):
    def __init__(self, null_display=_("NULL"), true_display=_("TRUE"), false_display=_("FALSE")):
        self.null_display = null_display
        self.true_display = true_display
        self.false_display = false_display
    
    def __call__(self, value):
        if value is None:
            return self.null_display
        if value:
            return self.true_display
        else:
            return self.false_display

def _get_field_value(item, field_setting, loop_index):
    field_name = field_setting["field"].split("__")[0]
    if field_name.lower() == "forloop.counter0":
        field_value = loop_index
    if field_name.lower() == "forloop.counter1":
        field_value = loop_index + 1
    else:
        if self.django_simple_export_admin_is_model_field(field_name):
            if hasattr(item, "get_{}_display".format(field_name)):
                field_value = getattr(item, "get_{}_display".format(field_name))()
            else:
                field_value = getattr(item, field_name, None)
                if field_name != field_setting["field"]:
                    for attr in field_setting["field"].split("__")[1:]:
                        if field_value:
                            field_value = getattr(field_value, attr, None)
        elif hasattr(item, field_name):
            field_value = getattr(item, field_name)()
        elif hasattr(self, field_name):
            field_value = getattr(self, field_name)(item)
        else:
            raise KeyError("field {0} not exists...".format(field_name))
        if "render" in field_setting:
            field_value = field_setting["render"](field_value)
        elif "empty_value" in field_setting and field_value is None:
            field_value = field_setting["empty_value"]
    return field_value
    

def do_export(items, fields, template=None, start_row_index=1, show_header=True):
    if template:
        workbook = get_workbook(template)
    else:
        workbook = openpyxl.Workbook()
    worksheet = workbook.active
    footer_values = {}
    row_index = start_row_index - 1
    # write header
    if show_header:
        for col_index, field_setting in enumerate(fields):
            label = field_setting["label"]
            worksheet.cell(row_index+1, col_index+1).value = label
        row_index += 1
    # write item list
    for loop_index, item in enumerate(items):
        for col_index, field_setting in enumerate(fields):
            value = self.django_simple_export_admin_get_field_value(item, field_setting, loop_index)
            real_col_index = field_setting.get("col", col_index+1)
            if not isinstance(real_col_index, (tuple, list, set)):
                worksheet.cell(row_index+1, real_col_index).value = value
            else:
                # split value into multi cells
                if isinstance(real_col_index, set):
                    real_col_index = list(real_col_index)
                    real_col_index.sort()
                cell_size = field_setting.get("cell_size", XLSX_CELL_SIZE_MAX)
                values = strutils.chunk(value, cell_size)
                for tmp_index in range(len(values)):
                    worksheet.cell(row_index+1, real_col_index[tmp_index]).value = values[tmp_index]
                # split value into multi cells end
            self.django_simple_export_admin_calc_footer_value(footer_values, col_index, field_setting, value)
        row_index += 1
    # write footer
    for col_index, field_setting in enumerate(fields):
        value = footer_values[col_index]
        if isinstance(value, Aggregate):
            value = value.final()
        if value:
            worksheet.cell(row_index+1, col_index+1).value = value
    row_index += 1
    return workbook

